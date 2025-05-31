import requests
from bs4 import BeautifulSoup
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

class GeorgianRecipeParser:
    def __init__(self, base_url, recipe_limit=8):
        self.base_url = base_url
        self.recipe_limit = recipe_limit
        self.recipes = []
    def fetch_recipe_links(self):
        r = requests.get(self.base_url)
        soup = BeautifulSoup(r.text, "html.parser")
        return soup.find_all("a", class_="h5")[:self.recipe_limit]
    def parse_ingredients(self, soup):
        ingredient_blocks = soup.find_all("div", class_="ingredient list-item")
        ingredients = []
        for block in ingredient_blocks:
            name_tag = block.find("a", class_="name")
            name = name_tag.text.strip() if name_tag else ""
            value_tag = block.find("span", class_="squant value")
            value = value_tag.text.strip() if value_tag else ""
            unit_tag = block.find("select", class_="recalc_s_num")
            unit_option = unit_tag.find("option", selected=True) if unit_tag else None
            unit = unit_option.text.strip() if unit_option else ""
            ingredients.append({
                "name": name,
                "amount": {
                    "value": value,
                    "unit": unit
                }
            })
        return ingredients
    def parse_recipe(self, link):
        title = link.text.strip()
        href = link.get("href")
        full_url = "https://1000.menu" + href
        ingredient_response = requests.get(full_url)
        ingredient_soup = BeautifulSoup(ingredient_response.text, "html.parser")
        ingredients = self.parse_ingredients(ingredient_soup)
        calories_tag = ingredient_soup.find("span", id="nutr_kcal")
        calories = calories_tag.text.strip() if calories_tag else None
        return {
            "title": title,
            "url": full_url,
            "calories": calories,
            "ingredients": ingredients
        }
    def collect_recipes(self):
        links = self.fetch_recipe_links()
        for link in links:
            recipe = self.parse_recipe(link)
            self.recipes.append(recipe)
            time.sleep(1)
    def save_to_excel(self, filename):
        rows = []
        for recipe in self.recipes:
            ingredients_str = ""
            for ing in recipe["ingredients"]:
                if len(ingredients_str) > 0:
                    ingredients_str += ", \n"
                name = str(ing["name"])
                value = str(ing["amount"]["value"])
                unit = str(ing["amount"]["unit"])
                ingredients_str += f"{name}: {value}{unit}"
            rows.append({
                "Наименование": recipe["title"],
                "Ссылка": recipe["url"],
                "Ккал": recipe["calories"],
                "Ингридиенты": ingredients_str
            })
        df = pd.DataFrame(rows)
        df.to_excel(filename, index=False)
        self.format_excel(filename)
    def format_excel(self, path):
        wb = load_workbook(path)
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            ws.column_dimensions[col_letter].width = max_length + 4
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(path)
def main():
    parser = GeorgianRecipeParser("https://1000.menu/catalog/gruzinskaya-kuxnya")
    parser.collect_recipes()
    parser.save_to_excel("georgian_recipes.xlsx")
if __name__ == "__main__":
    main()
